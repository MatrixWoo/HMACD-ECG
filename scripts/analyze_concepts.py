"""
P0 Experiments: Concept Intervention + Redundancy Analysis.

1. Concept Intervention (PCS):
   - Suppress each concept z_k = 0, measure prediction change
   - PCS_{k,c} = mean Δ prob for class c when concept k is suppressed
   - Output: PCS matrix + per-class top concepts + C31 special analysis

2. Redundancy Analysis:
   - Jaccard overlap between top-K activating samples per concept
   - Activation correlation matrix (z_i vs z_j)
   - Peak position distribution per concept (boundary artifact check)
   - Identifies duplicate concepts and suspected spurious concepts

Usage:
    python scripts/analyze_concepts.py \
        --ckpt results/hmacd_k32_seed42_best.pt \
        --device cuda:0
"""

import os, sys, argparse, time, itertools
import numpy as np
import torch
import torch.nn.functional as F
from torch.utils.data import DataLoader

sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

from src.models.resnet1d import ResNet1D
from src.models.concept_bank import HMACDModel
from src.datasets.ptbxl_dataset import PTBXLDataset, SUPERCLASS_LIST

LEAD_NAMES = ["I","II","III","aVR","aVL","aVF","V1","V2","V3","V4","V5","V6"]


def load_model(ckpt_path, device):
    ckpt = torch.load(ckpt_path, map_location=device)
    cfg = ckpt.get("config", {})
    K = cfg.get("model", {}).get("num_concepts", 32)
    backbone = ResNet1D(in_channels=12, num_classes=5)
    model = HMACDModel(backbone, num_concepts=K, num_classes=5)
    model.load_state_dict(ckpt["model_state_dict"])
    model.to(device).eval()
    return model, K


# ============================================================
#   Experiment 1: Concept Intervention (PCS)
# ============================================================

@torch.no_grad()
def compute_pcs_matrix(model, dataloader, K, device):
    """
    Suppress each concept, measure prediction change per class.
    Returns:
        pcs_matrix: [K, 5]  mean Δ prob (baseline - suppressed)
        pcs_std:    [K, 5]  std of Δ prob
        baseline_probs: [N, 5]
        all_probs:      [K, N, 5]  per-concept suppressed probs
    """
    model.eval()

    # ---- baseline pass ----
    baseline_probs = []
    all_labels = []
    for x, y in dataloader:
        x = x.to(device)
        logits, _, _ = model(x)
        probs = torch.sigmoid(logits).cpu()
        baseline_probs.append(probs)
        all_labels.append(y)
    baseline_probs = torch.cat(baseline_probs, dim=0)  # [N, 5]
    all_labels = torch.cat(all_labels, dim=0)           # [N, 5]
    N = baseline_probs.shape[0]

    print(f"  Baseline probs computed: {baseline_probs.shape}")
    print(f"  Now suppressing each of {K} concepts...")

    # ---- per-concept suppression ----
    all_probs = torch.zeros(K, N, 5)
    for k in range(K):
        probs_k = []
        for x, y in dataloader:
            x = x.to(device)
            logits, _, _ = model(x, intervene_concepts=[k])
            probs = torch.sigmoid(logits).cpu()
            probs_k.append(probs)
        all_probs[k] = torch.cat(probs_k, dim=0)

        if k % 8 == 0:
            print(f"    Concept {k:02d}/{K} done")

    # ---- PCS matrix ----
    pcs_matrix = torch.zeros(K, 5)
    pcs_std = torch.zeros(K, 5)
    pcs_pos = torch.zeros(K, 5)   # PCS on positive-label samples only

    for k in range(K):
        delta = baseline_probs - all_probs[k]  # [N, 5]
        pcs_matrix[k] = delta.mean(dim=0)
        pcs_std[k] = delta.std(dim=0)
        # positive-label only
        for c in range(5):
            mask = all_labels[:, c] > 0
            if mask.sum() > 0:
                pcs_pos[k, c] = delta[mask, c].mean()

    return pcs_matrix.numpy(), pcs_std.numpy(), pcs_pos.numpy(), baseline_probs.numpy(), all_labels.numpy()


# ============================================================
#   Experiment 2: Redundancy Analysis
# ============================================================

def compute_redundancy_metrics(all_z_np, top_k=50):
    """
    Compute concept redundancy metrics from activation matrix.

    Args:
        all_z_np: [N, K] concept activations across N samples

    Returns:
        jaccard:    [K, K] Jaccard overlap of top-K activating samples
        z_corr:     [K, K] Pearson r between concept activations
        top_indices: [K, top_k] sample indices per concept
        peak_stats:  dict with per-concept peak position stats
    """
    N, K = all_z_np.shape

    # ---- top-K sample indices per concept ----
    top_indices = np.zeros((K, top_k), dtype=int)
    for k in range(K):
        top_indices[k] = np.argsort(all_z_np[:, k])[-top_k:][::-1]

    # ---- Jaccard overlap ----
    jaccard = np.zeros((K, K))
    for i in range(K):
        set_i = set(top_indices[i])
        for j in range(K):
            if i == j:
                jaccard[i, j] = 1.0
            else:
                set_j = set(top_indices[j])
                inter = len(set_i & set_j)
                union = len(set_i | set_j)
                jaccard[i, j] = inter / union if union > 0 else 0.0

    # ---- Activation correlation matrix ----
    z_corr = np.corrcoef(all_z_np.T)  # [K, K]

    return jaccard, z_corr, top_indices


def find_redundant_pairs(jaccard, z_corr, threshold_jaccard=0.30, threshold_corr=0.80):
    """Identify concept pairs that are likely redundant."""
    K = jaccard.shape[0]
    redundant_pairs = []
    for i in range(K):
        for j in range(i+1, K):
            if jaccard[i, j] >= threshold_jaccard or z_corr[i, j] >= threshold_corr:
                redundant_pairs.append({
                    "c1": i, "c2": j,
                    "jaccard": jaccard[i, j],
                    "z_corr": z_corr[i, j],
                })
    # sort by jaccard desc
    redundant_pairs.sort(key=lambda x: -x["jaccard"])
    return redundant_pairs


def compute_peak_position_stats(all_sim, signal_len=1000):
    """
    Per-concept peak similarity position distribution.
    Returns mean, std, and histogram of peak positions.
    """
    K = len(all_sim[0])
    all_peaks = {k: [] for k in range(K)}

    for sim_k_list in all_sim:  # list of [K, L] tensors
        if isinstance(sim_k_list, torch.Tensor):
            sim_k_list = sim_k_list.numpy()
        L = sim_k_list.shape[-1]
        down_factor = signal_len / L
        for k in range(K):
            peak_pos = np.argmax(sim_k_list[k])  # feature map position
            peak_time = peak_pos * down_factor + down_factor / 2  # map to signal time
            all_peaks[k].append(peak_time)

    peak_stats = {}
    for k in range(K):
        arr = np.array(all_peaks[k])
        peak_stats[k] = {
            "mean": arr.mean(),
            "std": arr.std(),
            "min": arr.min(),
            "max": arr.max(),
            "near_end": ((arr > signal_len * 0.90) | (arr < signal_len * 0.10)).mean(),
        }
    return peak_stats


# ============================================================
#   Report Generation
# ============================================================

def print_pcs_summary(pcs_pos, class_names):
    """Print PCS matrix summary."""
    K = pcs_pos.shape[0]
    print(f"\n{'='*70}")
    print("EXPERIMENT 1: Concept Intervention (PCS)")
    print(f"{'='*70}")
    print("\nPCS[k,c] = mean Δ prob for class c when concept k is suppressed (positive-label samples only)")
    print(f"\n{'Concept':>10s}", end="")
    for c, name in enumerate(class_names):
        print(f"  {name:>8s}", end="")
    print("  |  Top-associated class")
    print("-" * 70)

    top_per_class = {c: [] for c in range(5)}
    for k in range(K):
        print(f"  C{k:02d}      ", end="")
        for c in range(5):
            val = pcs_pos[k, c]
            marker = " *" if abs(val) > 0.003 else "  "
            print(f"  {val:>+7.4f}{marker}", end="")
            top_per_class[c].append((k, val))
        print()

    # top-5 per class by PCS
    print(f"\nTop-5 concepts per class (by PCS magnitude):")
    for c, name in enumerate(class_names):
        sorted_concepts = sorted(top_per_class[c], key=lambda x: -abs(x[1]))[:5]
        s = "  ".join(f"C{k:02d}:{v:+.4f}" for k, v in sorted_concepts)
        print(f"  {name:>6s}: {s}")


def print_redundancy_summary(jaccard, z_corr, class_names, threshold=0.25):
    """Print concept redundancy analysis."""
    K = jaccard.shape[0]
    print(f"\n{'='*70}")
    print("EXPERIMENT 2: Redundancy Analysis")
    print(f"{'='*70}")

    # find redundant pairs
    pairs = find_redundant_pairs(jaccard, z_corr,
                                 threshold_jaccard=threshold,
                                 threshold_corr=0.80)

    print(f"\nRedundant concept pairs (Jaccard > {threshold} or z_corr > 0.80):")
    print(f"{'Pairs':>10s}  {'Jaccard':>8s}  {'z_corr':>8s}  Verdict")
    print("-" * 55)
    if not pairs:
        print("  No highly redundant pairs found.")
    for p in pairs:
        verdict = "⚠️ HIGH" if p["jaccard"] > 0.35 else "~ moderate"
        print(f"  C{p['c1']:02d}–C{p['c2']:02d}   {p['jaccard']:>8.4f}  {p['z_corr']:>8.4f}  {verdict}")

    # specific pair analysis
    focus_pairs = [(3, 14), (18, 21), (0, 4), (29, 21), (29, 18)]
    print(f"\nFocused pair analysis:")
    print(f"{'Pairs':>10s}  {'Jaccard':>8s}  {'z_corr':>8s}")
    print("-" * 35)
    for i, j in focus_pairs:
        print(f"  C{i:02d}–C{j:02d}   {jaccard[i,j]:>8.4f}  {z_corr[i,j]:>8.4f}")


def print_peak_analysis(peak_stats, K, signal_len=1000):
    """Print peak position analysis — boundary artifact check."""
    print(f"\n{'='*70}")
    print("Peak Position Analysis (Boundary Artifact Check)")
    print(f"{'='*70}")
    print(f"{'Concept':>8s}  {'Mean pos':>8s}  {'Std':>8s}  {'Near end%':>10s}  Flag")
    print("-" * 55)

    suspicious = []
    for k in range(K):
        s = peak_stats[k]
        near_end_pct = s["near_end"] * 100
        flag = ""
        if near_end_pct > 30:
            flag = "⚠️ BOUNDARY"
            suspicious.append((k, near_end_pct, s["mean"]))
        elif near_end_pct > 20:
            flag = "~ borderline"

        print(f"  C{k:02d}      {s['mean']:>8.1f}  {s['std']:>8.1f}  {near_end_pct:>9.1f}%  {flag}")

    if suspicious:
        print(f"\n⚠️  Concepts with >30% peaks near signal boundaries (potential artifacts):")
        for k, pct, mean_pos in suspicious:
            print(f"    C{k:02d}: {pct:.1f}% near ends, mean position = {mean_pos:.0f}")


# ============================================================
#   Main
# ============================================================

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--ckpt", type=str, default="results/hmacd_k32_seed42_best.pt")
    parser.add_argument("--data_path", type=str, default="/home/wuzuoxu/Data/ECG/1.0.3/")
    parser.add_argument("--device", type=str, default="cuda:0")
    parser.add_argument("--top_k", type=int, default=50, help="K for top-K sample overlap")
    parser.add_argument("--batch_size", type=int, default=128)
    parser.add_argument("--output", type=str, default="results/concept_analysis_results.xlsx")
    args = parser.parse_args()

    device = torch.device(args.device if torch.cuda.is_available() else "cpu")
    print(f"Device: {device}")

    # ---- load model ----
    print(f"Loading model: {args.ckpt}")
    model, K = load_model(args.ckpt, device)
    print(f"Model loaded: K={K}")

    # ---- data ----
    ds = PTBXLDataset(args.data_path, split="test", fold=10)
    loader = DataLoader(ds, batch_size=args.batch_size, shuffle=False,
                        num_workers=4, pin_memory=True)
    print(f"Test samples: {len(ds)}")

    # ================================================================
    #   Experiment 1: PCS Matrix
    # ================================================================
    t0 = time.time()
    pcs_all, pcs_std, pcs_pos, baseline_probs, all_labels = compute_pcs_matrix(
        model, loader, K, device
    )
    t1 = time.time()
    print(f"\nPCS computation done in {t1-t0:.0f}s")

    print_pcs_summary(pcs_pos, SUPERCLASS_LIST)

    # ================================================================
    #   Experiment 1b: C31 Special Analysis
    # ================================================================
    print(f"\n{'='*70}")
    print("C31 SPECIAL ANALYSIS (Suspected Spurious Concept)")
    print(f"{'='*70}")
    # find samples where C31 has high PCS
    c31_pcs_mi = pcs_pos[31, 1]  # MI class PCS
    c31_pcs_norm = pcs_pos[31, 0]  # NORM class PCS
    print(f"  PCS(C31, MI):  {c31_pcs_mi:+.4f} (positive-label only)")
    print(f"  PCS(C31, NORM):{c31_pcs_norm:+.4f} (positive-label only)")

    # get z activations for C31
    all_z_list = []
    all_sim_list = []
    with torch.no_grad():
        for x, y in loader:
            x = x.to(device)
            _, z, sim = model(x)
            all_z_list.append(z.cpu())
            for i in range(x.shape[0]):
                all_sim_list.append(sim[i].cpu())

    all_z = torch.cat(all_z_list, dim=0).numpy()  # [N, K]
    z_c31 = all_z[:, 31]
    c31_top_idx = np.argsort(z_c31)[-20:][::-1]

    print(f"\n  C31 top-10 activating samples:")
    print(f"  {'Idx':>6s}  {'z_C31':>8s}  {'True Label':>20s}  {'Pred Probs (MI/NORM)':>25s}")
    print(f"  {'-'*65}")
    for idx in c31_top_idx[:10]:
        true_labels = [SUPERCLASS_LIST[c] for c in range(5) if all_labels[idx, c] > 0]
        mi_prob = baseline_probs[idx, 1]
        norm_prob = baseline_probs[idx, 0]
        print(f"  {idx:>6d}  {z_c31[idx]:>8.4f}  {','.join(true_labels):>20s}  "
              f"MI={mi_prob:.3f} NORM={norm_prob:.3f}")

    # peak position analysis for C31
    c31_peaks = [np.argmax(s[31].numpy()) for s in all_sim_list]
    L = all_sim_list[0].shape[-1]
    down_factor = 1000 / L
    c31_times = [p * down_factor for p in c31_peaks]
    print(f"\n  C31 peak position stats: mean={np.mean(c31_times):.0f}, "
          f"std={np.std(c31_times):.0f}, "
          f"near_end(>900): {sum(1 for t in c31_times if t > 900)/len(c31_times)*100:.1f}%")

    # ================================================================
    #   Experiment 2: Redundancy Analysis
    # ================================================================
    t0 = time.time()
    jaccard, z_corr, top_indices = compute_redundancy_metrics(all_z, top_k=args.top_k)

    # peak position stats
    peak_stats = compute_peak_position_stats(all_sim_list, signal_len=1000)
    t1 = time.time()
    print(f"\nRedundancy analysis done in {t1-t0:.0f}s")

    print_redundancy_summary(jaccard, z_corr, SUPERCLASS_LIST, threshold=0.20)
    print_peak_analysis(peak_stats, K)

    # ================================================================
    #   Save to xlsx
    # ================================================================
    save_xlsx(args.output, pcs_pos, pcs_all, jaccard, z_corr, top_indices,
              peak_stats, SUPERCLASS_LIST, K)

    print(f"\n{'='*70}")
    print(f"All results saved to {args.output}")
    print(f"{'='*70}")


def save_xlsx(output_path, pcs_pos, pcs_all, jaccard, z_corr, top_indices,
              peak_stats, class_names, K):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    hf = Font(bold=True, size=11)
    hfill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
    border = Border(left=Side('thin'), right=Side('thin'), top=Side('thin'), bottom=Side('thin'))
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

    def write_header(ws, row, headers):
        for c, h in enumerate(headers, 1):
            cell = ws.cell(row=row, column=c, value=h)
            cell.font = hf; cell.fill = hfill; cell.border = border

    def write_data(ws, row, data):
        for i, r in enumerate(data):
            for j, v in enumerate(r):
                ws.cell(row=row+i, column=1+j, value=v).border = border

    def auto_width(ws, ncols, w=14):
        for c in range(1, ncols+1):
            ws.column_dimensions[get_column_letter(c)].width = w

    # ---- Sheet 1: PCS Matrix ----
    ws1 = wb.active
    ws1.title = "1-PCS-Matrix"
    ws1['A1'] = "Concept Intervention: PCS Matrix (positive-label samples)"
    ws1['A1'].font = Font(bold=True, size=13)

    headers = ["Concept"] + [f"PCS_{c}" for c in class_names] + ["Max |PCS|", "Top Class"]
    write_header(ws1, 3, headers)
    rows = []
    for k in range(K):
        pcs_row = pcs_pos[k]
        max_c = np.argmax(np.abs(pcs_row))
        rows.append([f"C{k:02d}"] + [round(v, 5) for v in pcs_row] +
                    [round(pcs_row[max_c], 5), class_names[max_c]])
    write_data(ws1, 4, rows)
    auto_width(ws1, len(headers))

    # Highlight top-5 per class
    for c in range(5):
        top_k = np.argsort(-np.abs(pcs_pos[:, c]))[:5]
        for k in top_k:
            ws1.cell(row=4+k, column=2+c).fill = green_fill

    # ---- Sheet 2: Jaccard Overlap ----
    ws2 = wb.create_sheet("2-Jaccard-Overlap")
    ws2['A1'] = "Concept Redundancy: Jaccard Overlap (top-50 samples)"
    ws2['A1'].font = Font(bold=True, size=13)

    headers2 = ["Concept"] + [f"C{i:02d}" for i in range(K)]
    write_header(ws2, 3, headers2)
    rows2 = []
    for i in range(K):
        rows2.append([f"C{i:02d}"] + [round(jaccard[i,j], 4) for j in range(K)])
    write_data(ws2, 4, rows2)

    # highlight redundant pairs
    for i in range(K):
        for j in range(K):
            if i != j and jaccard[i,j] > 0.30:
                ws2.cell(row=4+i, column=2+j).fill = red_fill
            elif i != j and jaccard[i,j] > 0.20:
                ws2.cell(row=4+i, column=2+j).fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    auto_width(ws2, K+1, w=7)
    ws2.column_dimensions['A'].width = 10

    # ---- Sheet 3: Activation Correlation ----
    ws3 = wb.create_sheet("3-Activation-Correlation")
    ws3['A1'] = "Concept Redundancy: Activation Correlation (Pearson r)"
    ws3['A1'].font = Font(bold=True, size=13)
    write_header(ws3, 3, headers2)
    rows3 = []
    for i in range(K):
        rows3.append([f"C{i:02d}"] + [round(z_corr[i,j], 4) for j in range(K)])
    write_data(ws3, 4, rows3)
    for i in range(K):
        for j in range(K):
            if i != j and abs(z_corr[i,j]) > 0.80:
                ws3.cell(row=4+i, column=2+j).fill = red_fill
    auto_width(ws3, K+1, w=7)
    ws3.column_dimensions['A'].width = 10

    # ---- Sheet 4: Redundant Pairs Summary ----
    ws4 = wb.create_sheet("4-Redundant-Pairs")
    ws4['A1'] = "Highly Redundant Concept Pairs (Jaccard > 0.20 or |z_corr| > 0.80)"
    ws4['A1'].font = Font(bold=True, size=13)
    write_header(ws4, 3, ["Concept 1", "Concept 2", "Jaccard", "z_corr", "Verdict"])
    # find all redundant pairs
    pairs = []
    for i in range(K):
        for j in range(i+1, K):
            if jaccard[i,j] > 0.20 or abs(z_corr[i,j]) > 0.80:
                verdict = "HIGH" if jaccard[i,j] > 0.30 else "MODERATE"
                pairs.append((i, j, jaccard[i,j], z_corr[i,j], verdict))
    pairs.sort(key=lambda x: -x[2])
    write_data(ws4, 4, pairs)
    auto_width(ws4, 5)

    # ---- Sheet 5: Peak Position Analysis ----
    ws5 = wb.create_sheet("5-Peak-Positions")
    ws5['A1'] = "Peak Position Analysis — Boundary Artifact Check"
    ws5['A1'].font = Font(bold=True, size=13)
    write_header(ws5, 3, ["Concept", "Mean Pos", "Std", "Min", "Max", "% Near End", "Flag"])
    rows5 = []
    for k in range(K):
        s = peak_stats[k]
        near_pct = s["near_end"] * 100
        flag = "⚠️ BOUNDARY" if near_pct > 30 else ("~ borderline" if near_pct > 20 else "")
        rows5.append([f"C{k:02d}", round(s['mean'],1), round(s['std'],1),
                      round(s['min'],1), round(s['max'],1), round(near_pct,1), flag])
    write_data(ws5, 4, rows5)
    for i, r in enumerate(rows5):
        if "BOUNDARY" in str(r[-1]):
            for c in range(1, 8):
                ws5.cell(row=4+i, column=c).fill = red_fill
    auto_width(ws5, 7)

    # ---- Sheet 6: PCS std (uncertainty) ----
    ws6 = wb.create_sheet("6-PCS-Detail")
    ws6['A1'] = "PCS Detail: mean ± std per concept per class (all samples)"
    ws6['A1'].font = Font(bold=True, size=13)
    write_header(ws6, 3, ["Concept"] + [f"{c}" for c in class_names])
    rows6 = []
    for k in range(K):
        rows6.append([f"C{k:02d}"] + [f"{pcs_all[k,c]:+.5f}" for c in range(5)])
    write_data(ws6, 4, rows6)
    auto_width(ws6, 6)

    os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)
    wb.save(output_path)


if __name__ == "__main__":
    main()
