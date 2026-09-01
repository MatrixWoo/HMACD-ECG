"""
Human Alignment Score (HAS) — Experiment 3 CORE.

Compute HAS_k = R²(z_k, g(U)) for each concept k:
- z_k: model concept activation
- U: PTB-XL+ human-engineered ECG features [2025-dim]
- g: regression model (Linear, Ridge, RandomForest, XGBoost)

Outputs:
- HAS table: per-concept R² scores
- HAS vs PCS scatter plot (core Figure 3)
- Four-quadrant classification of all 32 concepts

Usage:
    python scripts/compute_has.py --ckpt results/hmacd_k32_seed42_best.pt
"""

import os, sys, argparse, time, warnings
import numpy as np
import pandas as pd

import torch
from torch.utils.data import DataLoader

sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

from src.datasets.ptbxl_dataset import PTBXLDataset, SUPERCLASS_LIST
from src.datasets.ptbxl_plus_dataset import PTBXLPlusDataset
from src.models.resnet1d import ResNet1D
from src.models.concept_bank import HMACDModel

from sklearn.linear_model import LinearRegression, RidgeCV
from sklearn.ensemble import RandomForestRegressor
from sklearn.preprocessing import StandardScaler
from sklearn.metrics import r2_score
from joblib import Parallel, delayed

warnings.filterwarnings("ignore")


# ============================================================
#   Data loading with alignment
# ============================================================

def extract_concept_activations(model, dataloader, device):
    """Extract z [N, K] and ecg_ids from HMACD model."""
    model.eval()
    all_z = []
    with torch.no_grad():
        for x, _ in dataloader:
            x = x.to(device)
            _, z, _ = model(x)
            all_z.append(z.cpu())
    return torch.cat(all_z, dim=0).numpy()


def load_aligned_data(model, data_path, plus_path, device, batch_size=256):
    """
    Load concept activations z and human features U, aligned by ecg_id.
    Both datasets use the same strat_fold split → rows are aligned.
    We just verify alignment by ecg_id.
    """
    results = {}
    splits = {"train": (10, 9), "val": (9, 10), "test": (10, 10)}

    for split, (fold, val_fold) in splits.items():
        ds_ecg = PTBXLDataset(data_path, split=split, fold=fold)
        ds_feat = PTBXLPlusDataset(data_path, plus_path, split=split, fold=fold)

        # Get ecg_ids from both datasets (in order)
        ecg_ids_ecg = ds_ecg.df["ecg_id"].tolist()

        feat_ecg_ids = []
        for i in range(len(ds_feat)):
            row_idx = ds_feat.indices[i]
            eid = ds_feat._db.iloc[row_idx]["ecg_id"]
            feat_ecg_ids.append(int(eid))

        # Build map: ecg_id → position in feature array
        feat_pos_by_id = {eid: pos for pos, eid in enumerate(feat_ecg_ids)}

        # Extract concept activations z (same order as ds_ecg)
        loader = DataLoader(ds_ecg, batch_size=batch_size, shuffle=False,
                           num_workers=4, pin_memory=True)
        z_all = extract_concept_activations(model, loader, device)  # [N_ecg, K]

        # Collect labels + U
        y_all = []
        for _, y in loader:
            y_all.append(y)
        y_all = torch.cat(y_all, dim=0).numpy()

        # Build feature matrix from dataset
        U_all = np.zeros((len(ds_feat), ds_feat.M), dtype=np.float32)
        for i in range(len(ds_feat)):
            U_all[i] = ds_feat[i].numpy()

        # Align: keep only ecg_ids present in BOTH datasets, preserving order
        feat_id_set = set(feat_ecg_ids)
        z_aligned_list, U_aligned_list, y_aligned_list = [], [], []

        for ecg_pos, eid in enumerate(ecg_ids_ecg):
            eid_int = int(eid)
            if eid_int in feat_id_set:
                feat_pos = feat_pos_by_id[eid_int]
                z_aligned_list.append(z_all[ecg_pos])
                U_aligned_list.append(U_all[feat_pos])
                y_aligned_list.append(y_all[ecg_pos])

        z_aligned = np.array(z_aligned_list)
        U_aligned = np.array(U_aligned_list)
        y_aligned = np.array(y_aligned_list)

        results[split] = {
            "z": z_aligned,
            "U": U_aligned,
            "y": y_aligned,
            "n": z_aligned.shape[0],
        }

    return results


# ============================================================
#   HAS computation
# ============================================================

def compute_has_single(k, z_train, U_train, z_test, U_test):
    """Compute HAS for a single concept k with 4 regressors."""
    has_k = {}

    # Standardize features
    scaler = StandardScaler()
    U_train_s = scaler.fit_transform(U_train)
    U_test_s = scaler.transform(U_test)

    y_train = z_train[:, k]
    y_test = z_test[:, k]

    # 1. Linear Regression
    lr = LinearRegression()
    lr.fit(U_train_s, y_train)
    has_k["Linear"] = r2_score(y_test, lr.predict(U_test_s))

    # 2. Ridge Regression
    ridge = RidgeCV(alphas=[0.1, 1.0, 10.0, 100.0], cv=3)
    ridge.fit(U_train_s, y_train)
    has_k["Ridge"] = r2_score(y_test, ridge.predict(U_test_s))

    # 3. Random Forest
    rf = RandomForestRegressor(n_estimators=100, max_depth=15,
                                min_samples_leaf=10, random_state=42, n_jobs=-1)
    rf.fit(U_train_s, y_train)
    has_k["RF"] = r2_score(y_test, rf.predict(U_test_s))

    # 4. XGBoost
    try:
        from xgboost import XGBRegressor
        xgb = XGBRegressor(n_estimators=100, max_depth=6, learning_rate=0.1,
                          random_state=42, verbosity=0, n_jobs=1)
        xgb.fit(U_train_s, y_train)
        has_k["XGB"] = r2_score(y_test, xgb.predict(U_test_s))
    except ImportError:
        has_k["XGB"] = float("nan")

    return k, has_k


def compute_has_all(z_train, U_train, z_test, U_test, K, n_jobs=-1):
    """Compute HAS for all concepts in parallel."""
    print(f"  Computing HAS for {K} concepts with {n_jobs} parallel jobs...")
    t0 = time.time()

    results = Parallel(n_jobs=n_jobs, verbose=5)(
        delayed(compute_has_single)(k, z_train, U_train, z_test, U_test)
        for k in range(K)
    )

    # results is list of (k, has_k) tuples
    has_matrix = {}
    for k, has_k in results:
        has_matrix[k] = has_k

    t1 = time.time()
    print(f"  Done in {t1-t0:.0f}s")
    return has_matrix


# ============================================================
#   Visualization
# ============================================================

def plot_has_vs_pcs(has_values, pcs_values, save_path, class_names=SUPERCLASS_LIST):
    """HAS vs PCS scatter plot — core Figure 3."""
    import matplotlib
    matplotlib.use("Agg")
    import matplotlib.pyplot as plt
    from matplotlib.lines import Line2D

    K = len(has_values)

    # Color by dominant class
    dominant_class = np.argmax(np.abs(pcs_values), axis=1)  # [K]
    colors = plt.cm.tab10(np.linspace(0, 1, 5))

    fig, axes = plt.subplots(1, 5, figsize=(28, 6),
                              gridspec_kw={"width_ratios": [4, 4, 4, 4, 4]})

    for c in range(5):
        ax = axes[c]
        mask = dominant_class == c
        for k in range(K):
            if mask[k]:
                ax.scatter(has_values[k], pcs_values[k, c], s=80,
                          color=colors[c], edgecolors="black", linewidth=0.5,
                          zorder=3)
                ax.annotate(f"C{k:02d}", (has_values[k], pcs_values[k, c]),
                           fontsize=6, ha="center", va="bottom",
                           xytext=(0, 4), textcoords="offset points")

        ax.axhline(0, color="gray", linestyle="--", linewidth=0.5)
        ax.set_xlabel("HAS (R²)", fontsize=11)
        ax.set_ylabel(f"PCS ({class_names[c]})", fontsize=11)
        ax.set_title(f"{class_names[c]}", fontsize=13, fontweight="bold", color=colors[c])
        ax.grid(True, alpha=0.3)

    fig.suptitle("HAS vs PCS: Human Alignment × Predictive Contribution",
                 fontsize=15, fontweight="bold")
    fig.tight_layout()
    fig.savefig(save_path, dpi=150, bbox_inches="tight")
    plt.close(fig)
    return save_path


def plot_quadrant_matrix(has_values, pcs_max, save_path, class_names=SUPERCLASS_LIST):
    """Four-quadrant plot: HAS vs max |PCS|."""
    import matplotlib
    matplotlib.use("Agg")
    import matplotlib.pyplot as plt

    K = len(has_values)
    median_has = np.median(has_values)
    median_pcs = np.median(np.abs(pcs_max))

    fig, ax = plt.subplots(figsize=(12, 10))

    for k in range(K):
        has = has_values[k]
        pcs = pcs_max[k]
        dom_cls = np.argmax(np.abs(pcs_max[k])) if isinstance(pcs_max, np.ndarray) else 0

        ax.scatter(has, pcs, s=150, zorder=5, edgecolors="black", linewidth=0.5)
        ax.annotate(f"C{k:02d}", (has, pcs), fontsize=8, ha="center", va="bottom",
                   xytext=(0, 6), textcoords="offset points")

    # Quadrant lines
    ax.axhline(median_pcs, color="gray", linestyle="--", alpha=0.5)
    ax.axvline(median_has, color="gray", linestyle="--", alpha=0.5)

    # Quadrant labels
    ax.text(median_has + 0.03, median_pcs + 0.02, "Known Clinical\n(High HAS, High PCS)", ha="left", fontsize=9, color="green")
    ax.text(0.02, median_pcs + 0.02, "Model-Discovered\n(Low HAS, High PCS)", ha="left", fontsize=9, color="blue")
    ax.text(median_has + 0.03, 0.002, "Redundant Human\n(High HAS, Low PCS)", ha="left", fontsize=9, color="orange")
    ax.text(0.02, 0.002, "Spurious / Candidate\n(Low HAS, Low PCS)", ha="left", fontsize=9, color="red")

    ax.set_xlabel("HAS (R², Random Forest)", fontsize=12)
    ax.set_ylabel("max |PCS|", fontsize=12)
    ax.set_title("Human-Model Disagreement Matrix (32 Concepts)", fontsize=14, fontweight="bold")
    ax.grid(True, alpha=0.2)

    fig.tight_layout()
    fig.savefig(save_path, dpi=150, bbox_inches="tight")
    plt.close(fig)
    return save_path


# ============================================================
#   Main
# ============================================================

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--ckpt", type=str, default="results/hmacd_k32_seed42_best.pt")
    parser.add_argument("--data_path", type=str, default="/home/wuzuoxu/Data/ECG/1.0.3/")
    parser.add_argument("--plus_path", type=str,
                        default="/home/wuzuoxu/Data/ECG/ptb-xl-plus/1.0.0/features/")
    parser.add_argument("--device", type=str, default="cuda:0")
    parser.add_argument("--batch_size", type=int, default=256)
    parser.add_argument("--n_jobs", type=int, default=8)
    parser.add_argument("--output", type=str, default="results/has_results.xlsx")
    args = parser.parse_args()

    device = torch.device(args.device if torch.cuda.is_available() else "cpu")
    print(f"Device: {device}")

    # ---- Load model ----
    print(f"Loading model: {args.ckpt}")
    ckpt = torch.load(args.ckpt, map_location=device)
    cfg = ckpt.get("config", {})
    K = cfg.get("model", {}).get("num_concepts", 32)
    backbone = ResNet1D(in_channels=12, num_classes=5)
    model = HMACDModel(backbone, num_concepts=K, num_classes=5)
    model.load_state_dict(ckpt["model_state_dict"])
    model.to(device).eval()
    print(f"Model loaded: K={K}")

    # ---- Load aligned data ----
    print("Loading aligned concept activations and human features...")
    data = load_aligned_data(model, args.data_path, args.plus_path, device, args.batch_size)
    for split, d in data.items():
        print(f"  {split}: z={d['z'].shape}, U={d['U'].shape}, y={d['y'].shape}")

    z_train, U_train = data["train"]["z"], data["train"]["U"]
    z_test, U_test = data["test"]["z"], data["test"]["U"]

    # ---- Compute HAS ----
    print("\n" + "="*60)
    print("Computing Human Alignment Scores")
    print("="*60)
    has_matrix = compute_has_all(z_train, U_train, z_test, U_test, K, n_jobs=args.n_jobs)

    # ---- Print HAS table ----
    print(f"\n{'Concept':>8s}  {'Linear':>8s}  {'Ridge':>8s}  {'RF':>8s}  {'XGB':>8s}  |  {'Verdict':>25s}")
    print("-" * 80)
    has_rf = np.array([has_matrix[k]["RF"] for k in range(K)])
    median_has = np.median(has_rf)
    for k in range(K):
        h = has_matrix[k]
        rf_val = h["RF"]
        verdict = "High alignment" if rf_val > 0.3 else ("Moderate" if rf_val > 0.1 else "Low / Novel")
        print(f"  C{k:02d}      {h['Linear']:>8.4f}  {h['Ridge']:>8.4f}  {rf_val:>8.4f}  {h['XGB']:>8.4f}  |  {verdict}")

    print(f"\nHAS(RF) statistics: mean={has_rf.mean():.3f}, median={median_has:.3f}, "
          f"min={has_rf.min():.3f}, max={has_rf.max():.3f}")
    print(f"High HAS (>0.3): {(has_rf > 0.3).sum()}/32  concepts explained by traditional features")
    print(f"Low HAS (<0.1): {(has_rf < 0.1).sum()}/32  concepts NOT explained by traditional features")

    # ---- Load PCS results ----
    pcs_pos = np.zeros((K, 5))
    pcs_xlsx = "results/concept_analysis_results.xlsx"
    if os.path.exists(pcs_xlsx):
        try:
            # The xlsx has title+header rows; skip to raw data
            pcs_raw = pd.read_excel(pcs_xlsx, sheet_name="1-PCS-Matrix",
                                     header=None, skiprows=3)
            # pcs_raw: columns = [Concept, PCS_NORM, PCS_MI, PCS_STTC, PCS_CD, PCS_HYP, Max|PCS|, TopClass]
            for k in range(K):
                row = pcs_raw.iloc[k]
                for c in range(5):
                    pcs_pos[k, c] = float(row.iloc[1 + c])
            print(f"Loaded PCS matrix from {pcs_xlsx}")
        except Exception as e:
            print(f"Could not load PCS: {e}")
            pcs_pos = np.zeros((K, 5))
    else:
        print("PCS results not found, using zeros (run analyze_concepts.py first)")
        pcs_pos = np.zeros((K, 5))

    pcs_max = np.max(np.abs(pcs_pos), axis=1)  # [K]

    # ---- HAS vs PCS scatter ----
    os.makedirs("figures", exist_ok=True)
    f1 = plot_has_vs_pcs(has_rf, pcs_pos, "figures/has_vs_pcs.png")
    print(f"\nSaved: {f1}")

    f2 = plot_quadrant_matrix(has_rf, pcs_max, "figures/disagreement_matrix.png")
    print(f"Saved: {f2}")

    # ---- Classify concepts ----
    print(f"\n{'='*60}")
    print("Four-Quadrant Concept Classification")
    print(f"{'='*60}")
    median_pcs = np.median(pcs_max)
    print(f"  Median HAS(RF): {median_has:.3f}")
    print(f"  Median max|PCS|: {median_pcs:.4f}")

    categories = {"Known Clinical": [], "Model-Discovered": [],
                   "Redundant Human": [], "Spurious / Candidate": []}
    for k in range(K):
        has = has_rf[k]
        pcs = pcs_max[k]
        if has >= median_has and pcs >= median_pcs:
            cat = "Known Clinical"
        elif has < median_has and pcs >= median_pcs:
            cat = "Model-Discovered"
        elif has >= median_has and pcs < median_pcs:
            cat = "Redundant Human"
        else:
            cat = "Spurious / Candidate"
        categories[cat].append(k)

    for cat, concepts in categories.items():
        concept_str = ", ".join(f"C{k:02d}" for k in sorted(concepts))
        print(f"\n  {cat} ({len(concepts)} concepts):")
        print(f"    {concept_str}")

    # ---- Save xlsx ----
    save_has_xlsx(args.output, has_matrix, has_rf, pcs_pos, pcs_max, categories, K)

    print(f"\n{'='*60}")
    print(f"Results saved to {args.output}")
    print(f"Figures saved to figures/")
    print(f"{'='*60}")


def save_has_xlsx(output_path, has_matrix, has_rf, pcs_pos, pcs_max, categories, K):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    hf = Font(bold=True, size=11)
    hfill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
    border = Border(left=Side('thin'), right=Side('thin'), top=Side('thin'), bottom=Side('thin'))

    def write_header(ws, row, headers):
        for c, h in enumerate(headers, 1):
            cell = ws.cell(row=row, column=c, value=h)
            cell.font = hf; cell.fill = hfill; cell.border = border

    def write_data(ws, row, data):
        for i, r in enumerate(data):
            for j, v in enumerate(r):
                ws.cell(row=row+i, column=1+j, value=v).border = border

    # Sheet 1: HAS Summary
    ws1 = wb.active
    ws1.title = "1-HAS-Summary"
    ws1['A1'] = "Human Alignment Score (HAS) — R²(z_k, g(U))"
    ws1['A1'].font = Font(bold=True, size=13)
    headers = ["Concept", "HAS(Linear)", "HAS(Ridge)", "HAS(RF)", "HAS(XGB)",
               "max|PCS|", "Category"]
    write_header(ws1, 3, headers)
    rows = []
    for k in range(K):
        cat = ""
        for cname, clist in categories.items():
            if k in clist:
                cat = cname
                break
        rows.append([f"C{k:02d}", round(has_matrix[k]["Linear"],4),
                     round(has_matrix[k]["Ridge"],4), round(has_rf[k],4),
                     round(has_matrix[k]["XGB"],4), round(pcs_max[k],5), cat])
    write_data(ws1, 4, rows)

    # Sheet 2: HAS × PCS per class
    ws2 = wb.create_sheet("2-HAS-PCS-per-Class")
    ws2['A1'] = "HAS vs PCS per Disease Class"
    ws2['A1'].font = Font(bold=True, size=13)
    h2 = ["Concept"] + [f"HAS_{c}" for c in SUPERCLASS_LIST] + ["HAS(RF)"]
    # For now just put HAS(RF) for all classes (same)
    write_header(ws2, 3, h2)
    rows2 = []
    for k in range(K):
        rows2.append([f"C{k:02d}"] + [round(has_rf[k],4)]*5 + [round(has_rf[k],4)])
    write_data(ws2, 4, rows2)

    # Sheet 3: Four-Quadrant Classification
    ws3 = wb.create_sheet("3-Four-Quadrants")
    ws3['A1'] = "Four-Quadrant Concept Classification"
    ws3['A1'].font = Font(bold=True, size=13)
    write_header(ws3, 3, ["Category", "#Concepts", "Concepts"])
    rows3 = []
    for cat, concepts in categories.items():
        rows3.append([cat, len(concepts), ", ".join(f"C{k:02d}" for k in sorted(concepts))])
    write_data(ws3, 4, rows3)

    os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)
    wb.save(output_path)


if __name__ == "__main__":
    main()
