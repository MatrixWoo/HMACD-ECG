"""
Map each concept to top PTB-XL+ clinical features via HAS regression.

For each concept k:
  - Train RF regressor: U_train → z_train[:,k]
  - Extract top-10 feature importances
  - Map feature names to clinical descriptions
  - Output: per-concept clinical profile

Usage:
    python scripts/concept_feature_importance.py \
        --ckpt results/hmacd_k32_seed42_best.pt
"""

import os, sys, argparse, time, warnings
import numpy as np
import pandas as pd

import torch
from torch.utils.data import DataLoader

sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

from src.datasets.ptbxl_dataset import PTBXLDataset, SUPERCLASS_LIST
from src.datasets.ptbxl_plus_dataset import PTBXLPlusDataset
from src.models.resnet1d import ResNet1D
from src.models.concept_bank import HMACDModel

from sklearn.ensemble import RandomForestRegressor
from sklearn.preprocessing import StandardScaler
from joblib import Parallel, delayed

warnings.filterwarnings("ignore")


def load_aligned_data(model, data_path, plus_path, device, batch_size=256):
    """Load aligned z and U. Same as compute_has.py."""
    results = {}
    for split, fold, val_fold in [("train", 10, 9), ("test", 10, 10)]:
        ds_ecg = PTBXLDataset(data_path, split=split, fold=fold)
        ds_feat = PTBXLPlusDataset(data_path, plus_path, split=split, fold=fold)

        # Extract z
        loader = DataLoader(ds_ecg, batch_size=batch_size, shuffle=False,
                           num_workers=0, pin_memory=True)
        model.eval()
        all_z = []
        with torch.no_grad():
            for x, _ in loader:
                x = x.to(device)
                _, z, _ = model(x)
                all_z.append(z.cpu())
        z_all = torch.cat(all_z, dim=0).numpy()

        # Build U
        ecg_ids_ecg = ds_ecg.df["ecg_id"].tolist()
        feat_ecg_ids = []
        for i in range(len(ds_feat)):
            row_idx = ds_feat.indices[i]
            eid = ds_feat._db.iloc[row_idx]["ecg_id"]
            feat_ecg_ids.append(int(eid))
        feat_pos_by_id = {eid: pos for pos, eid in enumerate(feat_ecg_ids)}

        U_all = np.zeros((len(ds_feat), ds_feat.M), dtype=np.float32)
        for i in range(len(ds_feat)):
            U_all[i] = ds_feat[i].numpy()

        # Align
        feat_id_set = set(feat_ecg_ids)
        z_list, U_list = [], []
        for pos, eid in enumerate(ecg_ids_ecg):
            eid_int = int(eid)
            if eid_int in feat_id_set:
                z_list.append(z_all[pos])
                U_list.append(U_all[feat_pos_by_id[eid_int]])

        results[split] = {
            "z": np.array(z_list), "U": np.array(U_list),
        }
    return results


def compute_feature_importance_single(k, z_train, U_train, feature_names):
    """Train XGBoost and return top feature importances for concept k."""
    from xgboost import XGBRegressor
    # Subsample for speed: use 5000 random samples
    np.random.seed(42)
    idx = np.random.choice(len(z_train), min(5000, len(z_train)), replace=False)
    U_sub = U_train[idx]
    y_sub = z_train[idx, k]

    xgb = XGBRegressor(n_estimators=50, max_depth=6, learning_rate=0.1,
                       subsample=0.8, colsample_bytree=0.8,
                       random_state=42, verbosity=0, n_jobs=1)
    xgb.fit(U_sub, y_sub)

    # Get top-20 features by importance
    importances = xgb.feature_importances_
    top_idx = np.argsort(importances)[-20:][::-1]
    top_features = []
    for idx in top_idx:
        top_features.append({
            "rank": len(top_features) + 1,
            "feature": feature_names[idx],
            "importance": float(importances[idx]),
        })
    return top_features


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--ckpt", type=str, default="results/hmacd_k32_seed42_best.pt")
    parser.add_argument("--data_path", type=str, default="/home/wuzuoxu/Data/ECG/1.0.3/")
    parser.add_argument("--plus_path", type=str,
                        default="/home/wuzuoxu/Data/ECG/ptb-xl-plus/1.0.0/features/")
    parser.add_argument("--device", type=str, default="cuda:0")
    parser.add_argument("--n_jobs", type=int, default=16)
    args = parser.parse_args()

    device = torch.device(args.device if torch.cuda.is_available() else "cpu")

    # Load model
    ckpt = torch.load(args.ckpt, map_location=device)
    K = ckpt.get("config", {}).get("model", {}).get("num_concepts", 32)
    backbone = ResNet1D(in_channels=12, num_classes=5)
    model = HMACDModel(backbone, num_concepts=K, num_classes=5)
    model.load_state_dict(ckpt["model_state_dict"])
    model.to(device).eval()

    # Load data
    print("Loading aligned data...")
    data = load_aligned_data(model, args.data_path, args.plus_path, device)
    z_train, U_train = data["train"]["z"], data["train"]["U"]

    # Get feature names
    ds_feat = PTBXLPlusDataset(args.data_path, args.plus_path, split="train")
    feature_names = ds_feat.feature_names
    print(f"Train: z={z_train.shape}, U={U_train.shape}, features={len(feature_names)}")

    # Load feature descriptions
    desc_path = os.path.join(args.plus_path, "feature_description.csv")
    if os.path.exists(desc_path):
        desc_df = pd.read_csv(desc_path)
        # Build lookup: feature name → description
        # PTB-XL+ has 3 naming conventions; try all
        feat_desc = {}
        for _, row in desc_df.iterrows():
            for col in ["unig_feature", "12sl_feature", "ecgdeli_feature"]:
                name = str(row[col]) if pd.notna(row[col]) else ""
                if name and name != "nan":
                    feat_desc[name] = str(row["description"]) if pd.notna(row["description"]) else ""
    else:
        feat_desc = {}

    # Compute feature importance for all concepts
    # Compute feature importance for all concepts (sequential — XGBoost is fast)
    print(f"Computing feature importance for {K} concepts...")
    t0 = time.time()

    fi_by_concept = {}
    for k in range(K):
        top_features = compute_feature_importance_single(k, z_train, U_train, feature_names)
        fi_by_concept[k] = top_features
        if k % 8 == 0:
            t1 = time.time()
            print(f"  Concept {k:02d}/{K} done ({t1-t0:.0f}s)")

    t1 = time.time()
    print(f"Done in {t1-t0:.0f}s")

    # ---- Print results ----
    # Focus on key concepts: MI (C03, C14), STTC (C07, C21), Spurious (C31)
    focus = [3, 14, 7, 21, 0, 31]
    for k in focus:
        print(f"\n{'='*70}")
        print(f"Concept {k:02d} — Top-10 Clinical Features")
        print(f"{'='*70}")
        for f in fi_by_concept[k][:10]:
            desc = feat_desc.get(f["feature"], "")
            desc_str = f"  ← {desc}" if desc else ""
            print(f"  [{f['rank']:2d}] {f['feature']:<35s}  imp={f['importance']:.4f}{desc_str}")

    # ---- Save xlsx ----
    save_results(fi_by_concept, feat_desc, K)
    print(f"\nSaved to results/concept_feature_importance.xlsx")


def save_results(fi_by_concept, feat_desc, K):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Border, Side

    wb = Workbook()
    hf = Font(bold=True, size=11)
    hfill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
    border = Border(left=Side('thin'), right=Side('thin'), top=Side('thin'), bottom=Side('thin'))

    def write_header(ws, row, headers):
        for c, h in enumerate(headers, 1):
            cell = ws.cell(row=row, column=c, value=h)
            cell.font = hf; cell.fill = hfill; cell.border = border

    # Sheet 1: Summary — top-3 features per concept
    ws1 = wb.active
    ws1.title = "1-Summary"
    write_header(ws1, 1, ["Concept", "Top-1 Feature", "Importance", "Description",
                            "Top-2 Feature", "Importance", "Description",
                            "Top-3 Feature", "Importance", "Description"])
    for k in range(K):
        top = fi_by_concept[k]
        row = [f"C{k:02d}"]
        for i in range(3):
            f = top[i]
            row.extend([f["feature"], round(f["importance"], 5),
                       feat_desc.get(f["feature"], "")])
        for c, v in enumerate(row):
            ws1.cell(row=2+k, column=1+c, value=v).border = border

    # Sheet 2: Full — top-20 per concept
    ws2 = wb.create_sheet("2-Full")
    write_header(ws2, 1, ["Concept", "Rank", "Feature", "Importance", "Description"])
    row_idx = 2
    for k in range(K):
        for f in fi_by_concept[k]:
            for c, v in enumerate([f"C{k:02d}", f["rank"], f["feature"],
                                    round(f["importance"], 5),
                                    feat_desc.get(f["feature"], "")]):
                ws2.cell(row=row_idx, column=1+c, value=v).border = border
            row_idx += 1

    os.makedirs("results", exist_ok=True)
    wb.save("results/concept_feature_importance.xlsx")


if __name__ == "__main__":
    main()
